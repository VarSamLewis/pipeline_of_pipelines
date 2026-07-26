"""Auto-generated single-file Polars transformation pipeline.

Reads mapping.json from the same directory, loads source files from the
provided source folder, applies the approved mappings (including filters,
lookups, and aggregations), and writes one CSV per target table.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def load_mapping_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _catalog_tables(mapping_spec: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        table
        for catalog in mapping_spec.get("source_catalogs", [])
        for table in catalog.get("tables", [])
    ]


def _resolve_source_path(source_folder: Path, table: dict[str, Any]) -> Path:
    """Resolve one catalog table to exactly one content-verified local file."""
    original_filename = table.get("original_filename")
    raw_file_id = table.get("raw_file_id")
    candidates = []
    if original_filename:
        exact = source_folder / original_filename
        if exact.is_file():
            candidates.append(exact)
    if raw_file_id:
        candidates.extend(
            path
            for path in source_folder.glob(f"{raw_file_id}_*")
            if path.is_file() and path not in candidates
        )
    if len(candidates) != 1:
        raise ValueError(
            f"Expected exactly one local file for source table "
            f"{table['source_table_id']}, found {len(candidates)}"
        )
    path = candidates[0]
    expected_hash = table.get("file_sha256")
    if expected_hash:
        actual_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        if actual_hash != expected_hash:
            raise ValueError(
                f"Content hash mismatch for source table {table['source_table_id']}"
            )
    return path


def _catalog_column_names(table: dict[str, Any]) -> list[str]:
    return [column["normalized_name"] for column in table.get("columns", [])]


def _rows_to_dataframe(
    rows: list[list[Any]],
    table: dict[str, Any],
) -> pl.DataFrame:
    names = _catalog_column_names(table)
    width = len(names)
    header = [(str(value).strip() if value is not None else "") for value in rows[0]]
    data_rows = []
    for row in rows[1:]:
        padded = (list(row) + [None] * width)[:width]
        comparable = [
            str(value).strip() if value is not None else "" for value in padded
        ]
        if not any(value not in (None, "") for value in padded):
            continue
        if comparable == header[:width]:
            continue
        data_rows.append(padded)
    return pl.DataFrame(data_rows, schema=names, orient="row", infer_schema_length=None)


def _load_csv_table(path: Path, table: dict[str, Any]) -> pl.DataFrame:
    location = table["location"]
    encoding = location.get("encoding") or "utf-8"
    text = path.read_bytes().decode(encoding)
    rows = list(
        csv.reader(
            io.StringIO(text),
            delimiter=location.get("delimiter") or ",",
            quotechar=location.get("quote_char") or '"',
            strict=True,
        )
    )
    header_index = int(location["header_row"]) - 1
    return _rows_to_dataframe(rows[header_index:], table)


def _load_xlsx_table(path: Path, table: dict[str, Any]) -> pl.DataFrame:
    location = table["location"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[location["sheet_name"]]
        min_col, min_row, max_col, max_row = range_boundaries(
            location["cell_range"]
        )
        rows = [
            list(row)
            for row in worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        ]
    finally:
        workbook.close()
    return _rows_to_dataframe(rows, table)


def _load_legacy_source_dataframes(
    source_folder: Path,
) -> dict[str, pl.DataFrame]:
    """Keep old mapping artifacts executable during the catalog migration."""
    result = {}
    for path in sorted(source_folder.iterdir()):
        if path.suffix.lower() == ".csv":
            result[path.stem] = pl.read_csv(path)
        elif path.suffix.lower() in {".xlsx", ".xls"}:
            sheets = pl.read_excel(path, sheet_id=None, engine="openpyxl")
            if isinstance(sheets, dict):
                result.update({str(name): frame for name, frame in sheets.items()})
            else:
                result[path.stem] = sheets
    return result


def load_source_dataframes(
    mapping_spec: dict[str, Any],
    source_folder: Path,
) -> dict[str, pl.DataFrame]:
    """Load catalog tables by stable ID using their recorded parser settings."""
    tables = _catalog_tables(mapping_spec)
    if not tables:
        return _load_legacy_source_dataframes(source_folder)
    result = {}
    for table in tables:
        path = _resolve_source_path(source_folder, table)
        file_type = path.suffix.lower()
        if file_type == ".csv":
            frame = _load_csv_table(path, table)
        elif file_type == ".xlsx":
            frame = _load_xlsx_table(path, table)
        else:
            raise ValueError(
                f"Unsupported tabular source type for {table['source_table_id']}: "
                f"{file_type}"
            )
        result[table["source_table_id"]] = frame
    return result


def _concat_str(*args: Any) -> pl.Expr:
    """Concatenate a variadic list of strings/expressions with pl.concat_str."""
    exprs = [pl.lit(a) if not isinstance(a, pl.Expr) else a for a in args]
    return pl.concat_str(exprs)


def _eval_globals() -> dict[str, Any]:
    """Return the namespace used to evaluate polars_expression strings."""
    return {
        "pl": pl,
        "col": pl.col,
        "when": pl.when,
        "concat": _concat_str,
        "coalesce": pl.coalesce,
        "null": None,
        "Int64": pl.Int64,
        "Float64": pl.Float64,
        "String": pl.String,
        "Date": pl.Date,
        "Datetime": pl.Datetime,
        "Boolean": pl.Boolean,
    }


def apply_row_expression(
    df: pl.DataFrame,
    mapping: dict[str, Any],
) -> pl.DataFrame:
    """Apply a per-row Polars expression to a DataFrame."""
    target_column = mapping["target_column"]
    source_columns = mapping.get("source_columns", [])
    expression = mapping.get("polars_expression")

    if not source_columns:
        return df.with_columns(pl.lit(None).alias(target_column))

    first_source = source_columns[0]["source_column"]
    if expression:
        local_vars = {
            ref["source_column"]: df[ref["source_column"]]
            for ref in source_columns
            if ref["source_column"] in df.columns
        }
        try:
            result = eval(expression, _eval_globals(), local_vars)
            return df.with_columns(result.alias(target_column))
        except Exception:
            return df.with_columns(pl.lit(None).alias(target_column))
    if first_source in df.columns:
        return df.with_columns(pl.col(first_source).alias(target_column))
    return df.with_columns(pl.lit(None).alias(target_column))


def apply_filter(
    df: pl.DataFrame,
    mapping: dict[str, Any],
) -> pl.DataFrame:
    """Apply a Polars filter expression to a DataFrame."""
    expression = mapping.get("filter_expression")
    if not expression:
        return df
    local_vars = {name: df[name] for name in df.columns}
    try:
        mask = eval(expression, _eval_globals(), local_vars)
        return df.filter(mask)
    except Exception:
        return df


def apply_lookup(
    df: pl.DataFrame,
    mapping: dict[str, Any],
    source_dfs: dict[str, pl.DataFrame],
) -> pl.DataFrame:
    """Join a lookup table and bring in a value column."""
    target_column = mapping["target_column"]
    lookup_table = mapping.get("lookup_source_table")
    lookup_key = mapping.get("lookup_key")
    lookup_value = mapping.get("lookup_value")
    source_columns = mapping.get("source_columns", [])

    if not lookup_table:
        return df.with_columns(pl.lit(None).alias(target_column))
    if lookup_table not in source_dfs:
        lookup_table = Path(lookup_table).stem
    if lookup_table not in source_dfs:
        return df.with_columns(pl.lit(None).alias(target_column))
    if not lookup_key or not lookup_value:
        return df.with_columns(pl.lit(None).alias(target_column))

    left_key = source_columns[0]["source_column"] if source_columns else lookup_key
    # LLMs sometimes confuse the lookup-table column name with the source column;
    # fall back to case- and punctuation-insensitive matches on the current DataFrame.
    def _normalise(name: str) -> str:
        return name.lower().replace(" ", "_").replace("-", "_")

    if left_key not in df.columns:
        left_key = next(
            (
                c
                for c in df.columns
                if c.lower() == left_key.lower()
                or _normalise(c) == _normalise(left_key)
            ),
            left_key,
        )
    if left_key not in df.columns:
        return df.with_columns(pl.lit(None).alias(target_column))

    right = source_dfs[lookup_table].select([lookup_key, lookup_value]).unique()
    return df.join(right, left_on=left_key, right_on=lookup_key, how="left").rename(
        {lookup_value: target_column}
    )


def apply_aggregation(
    df: pl.DataFrame,
    mapping: dict[str, Any],
    source_dfs: dict[str, pl.DataFrame],
) -> pl.DataFrame:
    """Aggregate a source table and join the result back to the base table."""
    target_column = mapping["target_column"]
    agg_table = mapping.get("aggregation_source_table")
    agg_key = mapping.get("aggregation_group_key")
    agg_expression = mapping.get("aggregation_expression")
    source_columns = mapping.get("source_columns", [])

    if not agg_table or not agg_key or not agg_expression:
        return df.with_columns(pl.lit(None).alias(target_column))
    if agg_table not in source_dfs:
        agg_table = Path(agg_table).stem
    if agg_table not in source_dfs:
        return df.with_columns(pl.lit(None).alias(target_column))

    base_key = source_columns[0]["source_column"] if source_columns else agg_key

    agg_df = source_dfs[agg_table]
    local_vars = {name: agg_df[name] for name in agg_df.columns}
    try:
        agg_expr = eval(agg_expression, _eval_globals(), local_vars)
    except Exception:
        return df.with_columns(pl.lit(None).alias(target_column))

    grouped = agg_df.group_by(agg_key).agg(agg_expr.alias(target_column))
    return df.join(grouped, left_on=base_key, right_on=agg_key, how="left")


def apply_mapping(
    df: pl.DataFrame,
    mapping: dict[str, Any],
    source_dfs: dict[str, pl.DataFrame],
) -> pl.DataFrame:
    """Apply a single mapping to the working DataFrame."""
    ttype = mapping.get("transformation_type", "expression")
    if ttype == "filter":
        return apply_filter(df, mapping)
    if ttype == "lookup":
        return apply_lookup(df, mapping, source_dfs)
    if ttype == "aggregation":
        return apply_aggregation(df, mapping, source_dfs)
    return apply_row_expression(df, mapping)


def enforce_dtypes(
    df: pl.DataFrame,
    target_table: str,
    target_schema: dict[str, Any],
) -> pl.DataFrame:
    """Cast columns to the dtypes declared in the target schema."""
    table = next(
        (t for t in target_schema.get("tables", []) if t["name"] == target_table),
        None,
    )
    if table is None:
        return df
    casts = []
    for col in table.get("columns", []):
        dtype = col.get("dtype")
        name = col["name"]
        if not dtype or name not in df.columns:
            continue
        current_dtype = str(df[name].dtype).lower()
        if "null" in current_dtype:
            continue
        if dtype == "Date":
            if "date" in current_dtype and "datetime" not in current_dtype:
                continue
            if "datetime" in current_dtype:
                casts.append(pl.col(name).dt.date().alias(name))
            else:
                casts.append(pl.col(name).str.to_date(strict=False).alias(name))
        elif dtype == "Datetime":
            if "datetime" in current_dtype:
                continue
            casts.append(pl.col(name).str.to_datetime(strict=False).alias(name))
        else:
            polars_dtype = getattr(pl, dtype, None)
            if polars_dtype is not None:
                casts.append(pl.col(name).cast(polars_dtype, strict=False))
    if casts:
        df = df.with_columns(casts)
    return df


def build_target_tables(
    mapping_spec: dict[str, Any],
    source_folder: Path,
) -> dict[str, pl.DataFrame]:
    """Build all target tables from the source files and mappings."""
    source_dfs = load_source_dataframes(mapping_spec, source_folder)
    target_schema = mapping_spec.get("target_schema_json", {})

    columns_by_table: dict[str, list[dict[str, Any]]] = {}
    for col in mapping_spec["columns"]:
        columns_by_table.setdefault(col["target_table"], []).append(col)

    result: dict[str, pl.DataFrame] = {}
    for target_table, columns in columns_by_table.items():
        # Determine base source table from the first non-aggregation/lookup mapping
        base_source = target_table
        for col in columns:
            if col.get("source_columns"):
                source_ref = col["source_columns"][0]
                base_source = source_ref.get(
                    "source_table_id",
                    source_ref["source_table"],
                )
                break

        # Legacy artifacts may still identify a source by filename.
        source_key = base_source
        if source_key not in source_dfs:
            stem_key = Path(base_source).stem
            source_key = stem_key if stem_key in source_dfs else base_source
        if source_key not in source_dfs:
            raise ValueError(f"Source table not found: {base_source}")

        df = source_dfs[source_key].clone()

        # Apply filters first
        for col in columns:
            if col.get("transformation_type") == "filter":
                df = apply_mapping(df, col, source_dfs)

        # Apply remaining mappings
        for col in columns:
            if col.get("transformation_type") != "filter":
                df = apply_mapping(df, col, source_dfs)

        target_cols = [c["target_column"] for c in columns]
        df = df.select([c for c in target_cols if c in df.columns])
        df = enforce_dtypes(df, target_table, target_schema)
        result[target_table] = df

    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-folder", required=True, type=Path)
    parser.add_argument("--output-folder", default=".", type=Path)
    args = parser.parse_args()

    pipeline_dir = Path(__file__).parent
    mapping_spec = load_mapping_json(pipeline_dir / "mapping.json")
    tables = build_target_tables(mapping_spec, args.source_folder)

    args.output_folder.mkdir(parents=True, exist_ok=True)
    # Write every target table so multi-table specs are complete.
    for name, df in tables.items():
        df.write_csv(args.output_folder / f"{name}.csv")


if __name__ == "__main__":
    main()
