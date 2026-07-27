"""Auto-generated single-file Polars transformation pipeline."""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def load_mapping_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _catalog_tables(mapping_spec: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        table
        for catalog in mapping_spec.get("source_catalogs", [])
        for table in catalog.get("tables", [])
    ]


def _resolve_source_path(source_folder: Path, table: dict[str, Any]) -> Path:
    original_filename = table.get("original_filename")
    raw_file_id = table.get("raw_file_id")
    candidates = []
    if original_filename:
        exact = source_folder / original_filename
        if exact.is_file():
            candidates.append(exact)
    if raw_file_id:
        candidates.extend(
            p
            for p in source_folder.glob(f"{raw_file_id}_*")
            if p.is_file() and p not in candidates
        )
    if len(candidates) != 1:
        raise ValueError(
            f"Expected exactly one local file for source table "
            f"{table['source_table_id']}, found {len(candidates)}"
        )
    path = candidates[0]
    expected_hash = table.get("file_sha256")
    if expected_hash:
        actual_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        if actual_hash != expected_hash:
            raise ValueError(
                f"Content hash mismatch for source table {table['source_table_id']}"
            )
    return path


def _catalog_column_names(table: dict[str, Any]) -> list[str]:
    return [column["normalized_name"] for column in table.get("columns", [])]


def _rows_to_dataframe(
    rows: list[list[Any]],
    table: dict[str, Any],
) -> pl.DataFrame:
    names = _catalog_column_names(table)
    width = len(names)
    header = [(str(value).strip() if value is not None else "") for value in rows[0]]
    data_rows = []
    for row in rows[1:]:
        padded = (list(row) + [None] * width)[:width]
        comparable = [
            str(value).strip() if value is not None else "" for value in padded
        ]
        if not any(value not in (None, "") for value in padded):
            continue
        if comparable == header[:width]:
            continue
        data_rows.append(padded)
    return pl.DataFrame(data_rows, schema=names, orient="row", infer_schema_length=None)


def _load_csv_table(path: Path, table: dict[str, Any]) -> pl.DataFrame:
    location = table["location"]
    encoding = location.get("encoding") or "utf-8"
    text = path.read_bytes().decode(encoding)
    rows = list(
        csv.reader(
            io.StringIO(text),
            delimiter=location.get("delimiter") or ",",
            quotechar=location.get("quote_char") or '"',
            strict=True,
        )
    )
    header_index = int(location["header_row"]) - 1
    return _rows_to_dataframe(rows[header_index:], table)


def _load_xlsx_table(path: Path, table: dict[str, Any]) -> pl.DataFrame:
    location = table["location"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[location["sheet_name"]]
        min_col, min_row, max_col, max_row = range_boundaries(
            location["cell_range"]
        )
        rows = [
            list(row)
            for row in worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        ]
    finally:
        workbook.close()
    return _rows_to_dataframe(rows, table)


def _load_legacy_source_dataframes(
    source_folder: Path,
) -> dict[str, pl.DataFrame]:
    result = {}
    for path in sorted(source_folder.iterdir()):
        if path.suffix.lower() == ".csv":
            result[path.stem] = pl.read_csv(path)
        elif path.suffix.lower() in {".xlsx", ".xls"}:
            sheets = pl.read_excel(path, sheet_id=None, engine="openpyxl")
            if isinstance(sheets, dict):
                result.update({str(name): frame for name, frame in sheets.items()})
            else:
                result[path.stem] = sheets
    return result


def load_source_dataframes(
    mapping_spec: dict[str, Any],
    source_folder: Path,
) -> dict[str, pl.DataFrame]:
    tables = _catalog_tables(mapping_spec)
    if not tables:
        return _load_legacy_source_dataframes(source_folder)
    result = {}
    for table in tables:
        path = _resolve_source_path(source_folder, table)
        file_type = path.suffix.lower()
        if file_type == ".csv":
            frame = _load_csv_table(path, table)
        elif file_type == ".xlsx":
            frame = _load_xlsx_table(path, table)
        else:
            raise ValueError(
                f"Unsupported tabular source type for {table['source_table_id']}: "
                f"{file_type}"
            )
        result[table["source_table_id"]] = frame
    return result

_NS: dict[str, Any] = {
    "pl": pl,
    "col": pl.col,
    "when": pl.when,
    "concat": lambda *a: pl.concat_str([
        pl.lit(x) if not isinstance(x, pl.Expr) else x for x in a
    ]),
    "coalesce": pl.coalesce,
    "null": None,
    "Int64": pl.Int64,
    "Float64": pl.Float64,
    "String": pl.String,
    "Date": pl.Date,
    "Datetime": pl.Datetime,
    "Boolean": pl.Boolean,
}

_DTYPES: dict[str, dict[str, str]] = {
    "packaging_submission": {
        "submission_id": "String",
        "reporting_period": "String",
        "organisation_id": "String",
        "site_id": "String",
        "nation": "String",
        "packaging_material": "String",
        "packaging_class": "String",
        "activity": "String",
        "weight_tonnes": "Float64",
        "submission_date": "Date"
    }
}

_OUTPUT_DIR: Path = Path('.')

def _write_table(name: str, df: pl.DataFrame) -> None:
    _dtype = _DTYPES.get(name, {})
    for col_name, dtype_str in _dtype.items():
        if col_name not in df.columns:
            continue
        current = str(df[col_name].dtype).lower()
        if "null" in current:
            continue
        if dtype_str == "Date":
            if "date" in current and "datetime" not in current:
                continue
            if "datetime" in current:
                df = df.with_columns(
                    pl.col(col_name).dt.date().alias(col_name)
                )
            else:
                df = df.with_columns(
                    pl.col(col_name)
                    .str.to_date(strict=False)
                    .alias(col_name)
                )
        elif dtype_str == "Datetime":
            if "datetime" in current:
                continue
            df = df.with_columns(
                pl.col(col_name)
                .str.to_datetime(strict=False)
                .alias(col_name)
            )
        else:
            polars_dtype = getattr(pl, dtype_str, None)
            if polars_dtype is not None:
                df = df.with_columns(
                    pl.col(col_name)
                    .cast(polars_dtype, strict=False)
                )
    df.write_csv(_OUTPUT_DIR / f"{name}.csv")


def build_target_tables() -> None:
    # --- target table: packaging_submission ---
    _sk = 'dabded35-62d5-5931-88cf-e4df20061c11'
    if _sk not in source_dfs:
        _stem = Path(_sk).stem
        _sk = _stem if _stem in source_dfs else _sk
    df = source_dfs[_sk].clone()

    _local = {'packaging_type': df['packaging_type']}
    _expr = eval("col('packaging_type').str.split(' ').first()", _NS, _local)
    df = df.with_columns(_expr.alias('packaging_class'))

    _local = {'record_id': df['record_id'], 'period': df['period'], 'site_ref': df['site_ref'], 'packaging_material_code': df['packaging_material_code']}
    _expr = eval("concat(col('record_id'), '-', col('period'), '-', col('site_ref'), '-', col('packaging_material_code'))", _NS, _local)
    df = df.with_columns(_expr.alias('submission_id'))

    _local = {'period': df['period']}
    _expr = eval("coalesce(col('period').str.replace(r'(?i)Q', 'Q'), col('period'))", _NS, _local)
    df = df.with_columns(_expr.alias('reporting_period'))

    _local = {'org_code': df['org_code']}
    _expr = eval("col('org_code')", _NS, _local)
    df = df.with_columns(_expr.alias('organisation_id'))

    _local = {'site_ref': df['site_ref']}
    _expr = eval("col('site_ref')", _NS, _local)
    df = df.with_columns(_expr.alias('site_id'))

    _lut = (
        source_dfs['3365057d-e8d8-5b90-8e37-ff5529316b24']
        .select(['site_ref', 'nation'])
        .unique()
    )
    df = df.join(
        _lut,
        left_on='site_ref',
        right_on='site_ref',
        how="left",
    ).rename({'nation': 'nation'})

    _lut = (
        source_dfs['14469db7-ab0c-54c5-ba9e-1b5fe6e19bc3']
        .select(['packaging_material_code', 'ea_material_name'])
        .unique()
    )
    df = df.join(
        _lut,
        left_on='packaging_material_code',
        right_on='packaging_material_code',
        how="left",
    ).rename({'ea_material_name': 'packaging_material'})

    _local = {'activity_type': df['activity_type']}
    _expr = eval("when(col('activity_type').str.contains('supplied')).then(pl.lit('Supplied as goods')).otherwise(col('activity_type'))", _NS, _local)
    df = df.with_columns(_expr.alias('activity'))

    _local = {'weight': df['weight'], 'unit': df['unit']}
    _expr = eval("when(col('unit').str.contains('(?i)kg')).then(col('weight') / 1000).otherwise(col('weight'))", _NS, _local)
    df = df.with_columns(_expr.alias('weight_tonnes'))

    _local = {'record_date': df['record_date']}
    _expr = eval("coalesce(col('record_date').str.to_date('%d/%m/%Y', strict=False), col('record_date').str.to_date('%Y-%m-%d', strict=False), col('record_date').str.to_date('%d-%b-%Y', strict=False))", _NS, _local)
    df = df.with_columns(_expr.alias('submission_date'))

    df = df.select(['packaging_class', 'submission_id', 'reporting_period', 'organisation_id', 'site_id', 'nation', 'packaging_material', 'activity', 'weight_tonnes', 'submission_date'])
    _write_table('packaging_submission', df)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-folder", required=True, type=Path)
    parser.add_argument("--output-folder", default=".", type=Path)
    args = parser.parse_args()

    pipeline_dir = Path(__file__).parent
    mapping_spec = load_mapping_json(pipeline_dir / "mapping.json")
    source_dfs = load_source_dataframes(mapping_spec, args.source_folder)
    _OUTPUT_DIR = args.output_folder
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    build_target_tables()
