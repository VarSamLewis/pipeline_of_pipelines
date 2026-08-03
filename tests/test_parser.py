"""Unit tests for the parser and profiler module."""

from __future__ import annotations

from pathlib import Path

import pytest
from parser import (
    build_polars_from_mapping_source,
    discover_source_tables,
    parse_email_to_dict,
    parse_pdf_to_text,
    parse_text_document,
    profile_polars_dataframe,
    read_csv_to_polars,
    read_excel_to_polars,
    summarise_sheet,
)


def test_read_csv_to_polars_parses_rows() -> None:
    """CSV bytes should parse into a Polars DataFrame."""
    csv_bytes = b"id,name,score\n1,Alice,88.5\n2,Bob,92.0\n"
    df = read_csv_to_polars(csv_bytes)
    assert df.shape == (2, 3)
    assert df.columns == ["id", "name", "score"]


def test_read_excel_to_polars_parses_sheets(tmp_path: Path) -> None:
    """Excel bytes should parse into a Polars DataFrame."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    xlsx_path = tmp_path / "book.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "records"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    wb.save(xlsx_path)

    df = read_excel_to_polars(xlsx_path.read_bytes(), sheet_name="records")
    assert df.shape == (1, 2)
    assert df.columns == ["id", "name"]


def test_summarise_sheet_detects_headers(tmp_path: Path) -> None:
    """Spreadsheet summarisation should detect header row and columns."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    xlsx_path = tmp_path / "summary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    wb.save(xlsx_path)

    summary = summarise_sheet(xlsx_path.read_bytes(), sheet_name="data")
    assert summary["sheet_name"] == "data"
    assert summary["header_row"] == 1
    assert [c["header"] for c in summary["columns"]] == ["id", "name"]


def test_parse_pdf_to_text_extracts_text() -> None:
    """PDF text extraction should return the embedded text."""
    pytest.importorskip("fpdf")
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    pdf.cell(200, 10, text="Mapping instructions", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(200, 10, text="cust_id maps to customer_id", new_x="LMARGIN", new_y="NEXT")
    pdf_bytes = bytes(pdf.output())

    result = parse_pdf_to_text(pdf_bytes)
    full_text = result["full_text"]
    assert "Mapping instructions" in full_text
    assert "cust_id maps to customer_id" in full_text


def test_parse_email_to_dict_extracts_headers_and_body() -> None:
    """Email parsing should extract headers and body text."""
    eml_bytes = (
        b"From: sender@example.com\r\n"
        b"To: receiver@example.com\r\n"
        b"Subject: Rules\r\n\r\n"
        b"Validate all identifiers.\r\n"
    )
    result = parse_email_to_dict(eml_bytes)
    assert result["headers"]["Subject"] == "Rules"
    assert "Validate all identifiers" in result["body_text"]


def test_parse_text_document_reads_plain_text() -> None:
    """Text document parsing should return the UTF-8 contents."""
    result = parse_text_document(b"Hello world", "text/plain")
    assert result["text"] == "Hello world"


def test_profile_polars_dataframe_reports_statistics() -> None:
    """Profiling should report null counts and unique values."""
    pytest.importorskip("polars")
    import polars as pl

    df = pl.DataFrame(
        {
            "id": [1, 2, 3],
            "name": ["Alice", "Bob", None],
            "amount": [10.5, 20.0, 30.0],
        }
    )
    profile = profile_polars_dataframe(df, "records")
    by_name = {c["column"]: c for c in profile["columns"]}
    assert by_name["id"]["non_null_count"] == 3
    assert by_name["name"]["null_count"] == 1
    assert by_name["amount"]["unique_count"] == 3


def test_build_polars_from_mapping_source_csv() -> None:
    """CSV raw bytes should load as a Polars DataFrame for mapping sources."""
    csv_bytes = b"a,b\n1,hello\n2,world\n"
    df = build_polars_from_mapping_source(csv_bytes, "csv", "a.csv")
    assert df.shape == (2, 2)
    assert df.columns == ["a", "b"]


def test_discover_csv_dialect_encoding_and_profile() -> None:
    """CSV discovery should preserve parser decisions and profile its columns."""
    csv_bytes = (
        'Customer export\r\nid;name;note\r\n1;André;"uses; delimiter"\r\n2;Zoë;\r\n'
    ).encode("cp1252")

    catalog = discover_source_tables(
        csv_bytes,
        "csv",
        original_filename="renamed.csv",
    )

    assert len(catalog.tables) == 1
    table = catalog.tables[0]
    assert table.location.encoding == "cp1252"
    assert table.location.delimiter == ";"
    assert table.location.header_row == 2
    assert table.row_count == 2
    assert [column.normalized_name for column in table.columns] == [
        "id",
        "name",
        "note",
    ]
    assert table.columns[2].null_rate == 0.5
    assert any(warning.code == "encoding_fallback" for warning in catalog.warnings)
    assert any(warning.code == "preamble_detected" for warning in table.warnings)


def test_discover_csv_table_id_is_independent_of_filename() -> None:
    """Renaming an unchanged file must not change its source table identity."""
    contents = b"id,name\n1,Alice\n"
    first = discover_source_tables(contents, "csv", original_filename="first.csv")
    second = discover_source_tables(contents, "csv", original_filename="renamed.csv")

    assert first.tables[0].source_table_id == second.tables[0].source_table_id


def test_discover_xlsx_catalogues_every_sheet_and_region(tmp_path: Path) -> None:
    """Every worksheet and practical table region should receive a catalog entry."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "multi.xlsx"
    wb = Workbook()
    first = wb.active
    first.title = "Customers"
    first.append(["Customer report"])
    first.append(["id", "name"])
    first.append([1, "Alice"])
    first.append([2, "Bob"])
    second = wb.create_sheet("Lookups")
    second.append(["code", "label"])
    second.append(["A", "Alpha"])
    second.append([])
    second.append([])
    second.append(["country", "region"])
    second.append(["GB", "Europe"])
    wb.save(path)

    catalog = discover_source_tables(
        path.read_bytes(),
        "xlsx",
        original_filename=path.name,
    )

    assert {table.location.sheet_name for table in catalog.tables} == {
        "Customers",
        "Lookups",
    }
    assert len(catalog.tables) == 3
    customers = next(
        table for table in catalog.tables if table.location.sheet_name == "Customers"
    )
    assert customers.location.header_row == 2
    assert customers.row_count == 2
    assert [column.normalized_name for column in customers.columns] == ["id", "name"]


def test_discover_xlsx_tolerates_one_blank_row_inside_table(tmp_path: Path) -> None:
    """A single blank row should not split one logical table into two regions."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "blank-row.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["id", "name"])
    sheet.append([1, "Alice"])
    sheet.append([])
    sheet.append([2, "Bob"])
    workbook.save(path)

    catalog = discover_source_tables(path.read_bytes(), "xlsx")

    assert len(catalog.tables) == 1
    assert catalog.tables[0].row_count == 2
    assert catalog.tables[0].location.cell_range == "A1:B4"


def test_discover_csv_reports_malformed_quoting() -> None:
    """Malformed quoting should fail with an actionable diagnostic."""
    with pytest.raises(ValueError, match="Malformed CSV"):
        discover_source_tables(b'id,name\n1,"Alice\n', "csv")
