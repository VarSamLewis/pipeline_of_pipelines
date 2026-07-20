"""Generate a rich, multi-source client fixture for end-to-end testing.

This version produces thousands of rows so the LLM mapping and Polars pipeline
are tested on realistic volume. It also injects deliberate data-quality issues
that should be surfaced by the mapping rules and evidence.
"""

from __future__ import annotations

import json
import random
from datetime import datetime, timedelta
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook

# Realistic categorical data
FIRST_NAMES = [
    "Acme",
    "Globex",
    "Soylent",
    "Initech",
    "Hooli",
    "Massive",
    "Umbrella",
    "Stark",
    "Wayne",
    "Cyberdyne",
]
LAST_NAMES = [
    "Corp",
    "Inc",
    "LLC",
    "Ltd",
    "Group",
    "Holdings",
    "Enterprises",
    "Industries",
    "Systems",
    "Solutions",
]
REGIONS = ["NE", "SW", "NW", "SE"]
REGION_FULL = {
    "NE": "North-East",
    "SW": "South-West",
    "NW": "North-West",
    "SE": "South-East",
}
PRODUCTS = [
    ("SKU-A1", "Widget Alpha", "Hardware"),
    ("SKU-B2", "Beta Service", "Services"),
    ("SKU-C3", "Gamma Bolt", "Hardware"),
    ("SKU-D4", "Delta Drive", "Hardware"),
    ("SKU-E5", "Epsilon Edge", "Software"),
    ("SKU-F6", "Zeta Framework", "Services"),
    ("SKU-G7", "Eta Engine", "Hardware"),
    ("SKU-H8", "Theta Toolkit", "Software"),
]


def _pdf(text_lines: list[str], path: Path) -> None:
    """Write a simple PDF with one line per cell."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=11)
    for line in text_lines:
        pdf.cell(0, 8, text=line, new_x="LMARGIN", new_y="NEXT")
    pdf.output(str(path))


def _email(subject: str, body: str, path: Path) -> None:
    """Write a minimal .eml file."""
    content = (
        f"From: data-team@example.com\n"
        f"To: onboarding@example.com\n"
        f"Subject: {subject}\n\n"
        f"{body}\n"
    )
    path.write_text(content)


def _txt(lines: list[str], path: Path) -> None:
    """Write a plain text file."""
    path.write_text("\n".join(lines) + "\n")


def _generate_customers(n: int = 1000, seed: int = 42) -> list[dict[str, object]]:
    """Generate a list of customer rows with some deliberate dirty data."""
    rng = random.Random(seed)
    start_date = datetime(2023, 1, 1)
    customers: list[dict[str, object]] = []
    for i in range(1, n + 1):
        name = f"{rng.choice(FIRST_NAMES)} {rng.choice(LAST_NAMES)}"
        # Inject a known misspelling for customer 103 so the known-issues rule fires.
        if i == 103:
            name = "Soylent Corp"
        region = rng.choice(REGIONS)
        signup = start_date + timedelta(days=rng.randint(0, 365))
        customers.append(
            {
                "cust_id": i,
                "cust_name": name,
                "region_cd": region,
                "signup_dt": signup.strftime("%Y-%m-%d"),
            }
        )
    return customers


def _generate_orders(
    n: int = 10000,
    n_customers: int = 1000,
    seed: int = 42,
) -> list[dict[str, object]]:
    """Generate order line rows referencing the customers."""
    rng = random.Random(seed)
    start_date = datetime(2023, 1, 1)
    orders: list[dict[str, object]] = []
    for i in range(1, n + 1):
        sku, _, _ = rng.choice(PRODUCTS)
        qty = rng.randint(1, 20)
        # Inject a few negative quantities as a data-quality trap.
        if i % 997 == 0:
            qty = -qty
        unit_price = round(rng.uniform(9.99, 499.99), 2)
        order_dt = start_date + timedelta(days=rng.randint(0, 365))
        orders.append(
            {
                "order_id": 10000 + i,
                "cust_id": rng.randint(1, n_customers),
                "prod_sku": sku,
                "qty": qty,
                "unit_price": unit_price,
                "order_dt": order_dt.strftime("%Y-%m-%d"),
            }
        )
    # Inject a few test orders that should be excluded (order_id starts with 9999).
    for j in range(5):
        orders.append(
            {
                "order_id": 9999000 + j,
                "cust_id": rng.randint(1, n_customers),
                "prod_sku": rng.choice(PRODUCTS)[0],
                "qty": 1,
                "unit_price": 1.0,
                "order_dt": start_date.strftime("%Y-%m-%d"),
            }
        )
    return orders


def generate(
    folder: Path | None = None,
    n_customers: int = 1000,
    n_orders: int = 10000,
) -> None:
    """Create the complex multi-file fixture.

    Args:
        folder: Destination directory. Defaults to the script's parent.
        n_customers: Number of customer rows to generate.
        n_orders: Number of order line rows to generate.
    """
    folder = folder or Path(__file__).parent
    folder.mkdir(parents=True, exist_ok=True)

    customers = _generate_customers(n_customers)
    orders = _generate_orders(n_orders, n_customers)

    # ------------------------------------------------------------------
    # Multi-sheet Excel workbook
    # ------------------------------------------------------------------
    wb = Workbook()

    ws_cust = wb.active
    ws_cust.title = "customers"
    ws_cust.append(["cust_id", "cust_name", "region_cd", "signup_dt"])
    for row in customers:
        ws_cust.append(
            [
                row["cust_id"],
                row["cust_name"],
                row["region_cd"],
                row["signup_dt"],
            ]
        )

    ws_orders = wb.create_sheet("orders")
    ws_orders.append(
        ["order_id", "cust_id", "prod_sku", "qty", "unit_price", "order_dt"]
    )
    for row in orders:
        ws_orders.append(
            [
                row["order_id"],
                row["cust_id"],
                row["prod_sku"],
                row["qty"],
                row["unit_price"],
                row["order_dt"],
            ]
        )

    ws_products = wb.create_sheet("products")
    ws_products.append(["prod_sku", "prod_name", "category"])
    for sku, name, category in PRODUCTS:
        ws_products.append([sku, name, category])

    wb.save(folder / "master_data.xlsx")

    # ------------------------------------------------------------------
    # PDFs providing mapping and business context
    # ------------------------------------------------------------------
    _pdf(
        [
            "Global Retail Onboarding Guide",
            "",
            "Map cust_id to customer_id.",
            "Map cust_name to customer_name.",
            "region_cd should become region using the region reference table.",
            "signup_dt should be parsed as signup_date.",
            "Total revenue per customer is sum(qty * unit_price) from orders.",
        ],
        folder / "onboarding_guide.pdf",
    )

    _pdf(
        [
            "Region Reference",
            "",
            "NE means North-East.",
            "SW means South-West.",
            "NW means North-West.",
            "SE means South-East.",
        ],
        folder / "region_reference.pdf",
    )

    _pdf(
        [
            "Revenue Recognition Policy",
            "",
            "Revenue is calculated as quantity multiplied by unit_price.",
            "Round line totals to two decimal places.",
            "Negative quantities are not allowed and should be rejected.",
        ],
        folder / "revenue_policy.pdf",
    )

    _pdf(
        [
            "Product Catalogue Notes",
            "",
            "prod_sku is the canonical product identifier.",
            "prod_name should be mapped to product_name.",
            "category should be normalised to title case.",
        ],
        folder / "product_catalogue.pdf",
    )

    # ------------------------------------------------------------------
    # Emails with rules, issues and validation context
    # ------------------------------------------------------------------
    _email(
        "Customer Data Quality Rules",
        "Please ensure:\n"
        "- customer_id is unique and not null\n"
        "- customer_name is trimmed and title-cased\n"
        "- region is expanded to full region name\n"
        "- signup_date is a valid ISO date\n",
        folder / "rules_customers.eml",
    )

    _email(
        "Order Validation Rules",
        "For every order line:\n"
        "- order_id must not be null\n"
        "- qty must be a positive integer\n"
        "- unit_price must be a positive number\n"
        "- line_total = qty * unit_price\n",
        folder / "rules_orders.eml",
    )

    _email(
        "Region Code Mapping",
        "Use the following mapping for region_cd:\n"
        "NE -> North-East\n"
        "SW -> South-West\n"
        "NW -> North-West\n"
        "SE -> South-East\n",
        folder / "region_mapping.eml",
    )

    _email(
        "Urgent: Known Data Issues - March",
        "The March extract has a known issue where customer 103 name is misspelled.\n"
        "Correct 'Soylent Corp' to 'Soylent Corporation' before loading.\n"
        "Also exclude any test orders with order_id starting with 9999.\n",
        folder / "known_issues_march.eml",
    )

    _email(
        "Revenue Reporting Requirements",
        "Finance requires:\n"
        "- revenue rounded to 2 decimal places\n"
        "- one row per customer for the customer summary\n"
        "- one row per order line for the order detail\n",
        folder / "reporting_requirements.eml",
    )

    # ------------------------------------------------------------------
    # Text files with glossary and supplementary context
    # ------------------------------------------------------------------
    _txt(
        [
            "Business Glossary",
            "",
            "customer_id: unique identifier for a customer account.",
            "customer_name: official registered trading name of the customer.",
            "region: geographic sales region written as full words.",
            "signup_date: date the customer account was created.",
            "total_revenue: lifetime revenue for the customer.",
            "order_id: unique identifier for a sales transaction.",
            "line_total: revenue for a single order line.",
        ],
        folder / "business_glossary.txt",
    )

    _txt(
        [
            "Data Provenance Notes",
            "",
            "Customer master data comes from the CRM export.",
            "Order data comes from the ERP transaction log.",
            "Product data comes from the product catalogue system.",
            "PDFs were produced by the data governance team.",
        ],
        folder / "provenance.txt",
    )

    _txt(
        [
            "Contact and Escalation",
            "",
            "Data owner: Jane Doe (jane.doe@example.com)",
            "Engineering owner: John Smith (john.smith@example.com)",
            "Escalation: data-governance@example.com",
        ],
        folder / "contacts.txt",
    )

    # ------------------------------------------------------------------
    # Target schema
    # ------------------------------------------------------------------
    target_schema = {
        "client_code": "complexclient",
        "name": "default",
        "description": "Curated customer and order data for Global Retail onboarding",
        "tables": [
            {
                "name": "customers",
                "description": "Cleaned customer summary with lifetime revenue",
                "columns": [
                    {
                        "name": "customer_id",
                        "dtype": "Int64",
                        "required": True,
                        "unique": True,
                    },
                    {"name": "customer_name", "dtype": "String", "required": True},
                    {"name": "region", "dtype": "String", "required": False},
                    {"name": "signup_date", "dtype": "Date", "required": False},
                    {"name": "total_revenue", "dtype": "Float64", "required": False},
                ],
            },
            {
                "name": "orders",
                "description": "Cleaned order line detail",
                "columns": [
                    {"name": "order_id", "dtype": "Int64", "required": True},
                    {"name": "customer_id", "dtype": "Int64", "required": True},
                    {"name": "product_name", "dtype": "String", "required": False},
                    {"name": "category", "dtype": "String", "required": False},
                    {"name": "quantity", "dtype": "Int64", "required": False},
                    {"name": "unit_price", "dtype": "Float64", "required": False},
                    {"name": "order_date", "dtype": "Date", "required": False},
                    {"name": "line_total", "dtype": "Float64", "required": False},
                ],
            },
        ],
    }
    (folder / "target_schema.json").write_text(json.dumps(target_schema, indent=2))

    print(f"Generated {n_customers} customers and {n_orders} orders in {folder}")


if __name__ == "__main__":
    generate()
