"""Parsers and profilers for heterogeneous client files.

This module converts raw file bytes into structured facts and Polars DataFrames.
Supported inputs include spreadsheets (CSV, XLSX), PDFs, emails (EML), and
plain-text documents (TXT, MD, DOCX). All functions are pure: they take bytes
and return structured data without side effects.
"""

from __future__ import annotations

import csv
import email
import hashlib
import io
import re
import unicodedata
import uuid
from collections import Counter
from datetime import date, datetime
from typing import Any, cast

import polars as pl
from models import (
    ParseWarning,
    SourceCatalog,
    SourceColumn,
    SourceLocation,
    SourceTable,
)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_SOURCE_ID_NAMESPACE = uuid.UUID("9a216e98-08b3-46c8-b07d-79f06ee97a02")
_COMMON_DELIMITERS = ",;\t|"


def normalize_column_name(value: Any, ordinal: int) -> str:
    """Return a deterministic, identifier-friendly source column name."""
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    text = re.sub(r"[^0-9a-zA-Z]+", "_", text).strip("_")
    return text or f"column_{ordinal}"


def _unique_names(
    headers: list[Any],
) -> tuple[list[tuple[str, str]], list[ParseWarning]]:
    """Normalize headers and make duplicate normalized names unambiguous."""
    seen: Counter[str] = Counter()
    names: list[tuple[str, str]] = []
    warnings: list[ParseWarning] = []
    for ordinal, value in enumerate(headers, start=1):
        original = str(value).strip() if value is not None else ""
        normalized = normalize_column_name(value, ordinal)
        seen[normalized] += 1
        if seen[normalized] > 1:
            warnings.append(
                ParseWarning(
                    code="duplicate_header",
                    message=(
                        f"Duplicate column name {original or normalized!r} was renamed."
                    ),
                    location=f"column {ordinal}",
                )
            )
            normalized = f"{normalized}_{seen[normalized]}"
        names.append((original or f"Column {ordinal}", normalized))
    return names, warnings


def _infer_type(values: list[Any]) -> str:
    non_null = [value for value in values if value not in (None, "")]
    if not non_null:
        return "null"
    if all(isinstance(value, bool) for value in non_null):
        return "boolean"
    if all(
        isinstance(value, int) and not isinstance(value, bool) for value in non_null
    ):
        return "integer"
    if all(
        isinstance(value, (int, float)) and not isinstance(value, bool)
        for value in non_null
    ):
        return "number"
    if all(isinstance(value, datetime) for value in non_null):
        return "datetime"
    if all(isinstance(value, (date, datetime)) for value in non_null):
        return "date"
    return "string"


def _profile_table(
    *,
    file_sha256: str,
    raw_file_id: uuid.UUID | None,
    original_filename: str | None,
    display_name: str,
    location: SourceLocation,
    headers: list[Any],
    rows: list[list[Any]],
    confidence: float,
    warnings: list[ParseWarning] | None = None,
) -> SourceTable:
    names, header_warnings = _unique_names(headers)
    table_identity = "|".join(
        [
            file_sha256,
            location.sheet_name or "",
            location.cell_range or "",
            str(location.header_row),
            ",".join(normalized for _, normalized in names),
        ]
    )
    table_id = str(uuid.uuid5(_SOURCE_ID_NAMESPACE, table_identity))
    columns: list[SourceColumn] = []
    for index, (original, normalized) in enumerate(names):
        values = [row[index] if index < len(row) else None for row in rows]
        non_null = [value for value in values if value not in (None, "")]
        cardinality = len({str(value) for value in non_null})
        null_count = len(values) - len(non_null)
        uniqueness = cardinality / len(non_null) if non_null else 0.0
        completeness = len(non_null) / len(values) if values else 0.0
        key_score = uniqueness * completeness
        columns.append(
            SourceColumn(
                source_column_id=str(
                    uuid.uuid5(_SOURCE_ID_NAMESPACE, f"{table_id}|{index}|{normalized}")
                ),
                ordinal=index + 1,
                original_name=original,
                normalized_name=normalized,
                inferred_type=_infer_type(values),
                examples=[str(value) for value in non_null[:5]],
                null_count=null_count,
                null_rate=(null_count / len(values)) if values else 0.0,
                cardinality=cardinality,
                candidate_key_score=round(key_score, 4),
            )
        )
    return SourceTable(
        source_table_id=table_id,
        raw_file_id=raw_file_id,
        file_sha256=file_sha256,
        original_filename=original_filename,
        display_name=display_name,
        location=location,
        row_count=len(rows),
        columns=columns,
        confidence=confidence,
        warnings=[*(warnings or []), *header_warnings],
    )


def _detect_encoding(file_bytes: bytes) -> tuple[str, str, list[ParseWarning]]:
    warnings: list[ParseWarning] = []
    candidates: list[str] = []
    if file_bytes.startswith(b"\xef\xbb\xbf"):
        candidates.append("utf-8-sig")
    elif file_bytes.startswith((b"\xff\xfe", b"\xfe\xff")):
        candidates.append("utf-16")
    candidates.extend(["utf-8", "cp1252"])
    attempted: set[str] = set()
    for encoding in candidates:
        if encoding in attempted:
            continue
        attempted.add(encoding)
        try:
            text = file_bytes.decode(encoding)
            if encoding == "cp1252":
                warnings.append(
                    ParseWarning(
                        code="encoding_fallback",
                        message=(
                            "CSV was decoded using Windows-1252 after UTF-8 failed."
                        ),
                        details={"encoding": encoding},
                    )
                )
            return encoding, text, warnings
        except UnicodeDecodeError:
            continue
    raise ValueError(
        "CSV encoding could not be detected (tried UTF-8 and Windows-1252)"
    )


def _newline_style(text: str) -> str:
    if "\r\n" in text:
        return "CRLF"
    if "\r" in text:
        return "CR"
    return "LF"


def _detect_delimiter(text: str) -> tuple[str, list[ParseWarning]]:
    """Choose the common delimiter producing the most consistent multi-field rows."""
    best: tuple[int, int, str] | None = None
    for delimiter in _COMMON_DELIMITERS:
        try:
            rows = list(
                csv.reader(io.StringIO(text[:65536]), delimiter=delimiter, strict=True)
            )
        except csv.Error:
            continue
        widths = Counter(len(row) for row in rows if row)
        multi_field = [(count, width) for width, count in widths.items() if width > 1]
        if not multi_field:
            continue
        count, width = max(multi_field)
        score = (count, width, delimiter)
        if best is None or score[:2] > best[:2]:
            best = score
    if best is not None:
        return best[2], []
    return ",", [
        ParseWarning(
            code="dialect_fallback",
            message=(
                "CSV delimiter could not be determined; comma-separated defaults "
                "were used."
            ),
        )
    ]


def _discover_csv(
    file_bytes: bytes,
    *,
    file_sha256: str,
    raw_file_id: uuid.UUID | None,
    original_filename: str | None,
) -> SourceCatalog:
    encoding, text, warnings = _detect_encoding(file_bytes)
    delimiter, dialect_warnings = _detect_delimiter(text)
    warnings.extend(dialect_warnings)
    quote_char = '"'
    try:
        parsed_rows = [
            list(row)
            for row in csv.reader(
                io.StringIO(text),
                delimiter=delimiter,
                quotechar=quote_char,
                strict=True,
            )
        ]
    except csv.Error as exc:
        raise ValueError(f"Malformed CSV: {exc}") from exc
    non_blank = [row for row in parsed_rows if any(cell.strip() for cell in row)]
    if not non_blank:
        warnings.append(
            ParseWarning(code="empty_file", message="CSV contains no table data.")
        )
        return SourceCatalog(
            raw_file_id=raw_file_id,
            file_sha256=file_sha256,
            original_filename=original_filename,
            file_type="csv",
            warnings=warnings,
        )
    widths = Counter(len(row) for row in non_blank)
    width = widths.most_common(1)[0][0]
    ragged = [
        index + 1 for index, row in enumerate(parsed_rows) if row and len(row) != width
    ]
    table_warnings: list[ParseWarning] = []
    if ragged:
        table_warnings.append(
            ParseWarning(
                code="ragged_rows",
                message=f"{len(ragged)} CSV row(s) have a different field count.",
                details={"rows": ragged[:20], "expected_fields": width},
            )
        )
    padded_rows = [(row + [None] * width)[:width] for row in non_blank]
    candidate_count = min(10, len(padded_rows))
    scores = [
        _row_score(padded_rows[index], padded_rows[index + 1 : index + 6])
        for index in range(candidate_count)
    ]
    header_index = max(range(candidate_count), key=scores.__getitem__)
    headers = padded_rows[header_index]
    rows = [row for row in padded_rows[header_index + 1 :] if row != headers]
    if header_index:
        table_warnings.append(
            ParseWarning(
                code="preamble_detected",
                message=f"Detected {header_index} preamble/title row(s).",
            )
        )
    repeated_headers = sum(row == headers for row in padded_rows[header_index + 1 :])
    if repeated_headers:
        table_warnings.append(
            ParseWarning(
                code="repeated_header",
                message=f"Removed {repeated_headers} repeated header row(s).",
            )
        )
    location = SourceLocation(
        header_row=header_index + 1,
        data_start_row=header_index + 2,
        encoding=encoding,
        delimiter=delimiter,
        quote_char=quote_char,
        newline=_newline_style(text),
    )
    table = _profile_table(
        file_sha256=file_sha256,
        raw_file_id=raw_file_id,
        original_filename=original_filename,
        display_name=original_filename or "CSV table",
        location=location,
        headers=headers,
        rows=rows,
        confidence=0.95 if not table_warnings else 0.8,
        warnings=table_warnings,
    )
    return SourceCatalog(
        raw_file_id=raw_file_id,
        file_sha256=file_sha256,
        original_filename=original_filename,
        file_type="csv",
        tables=[table],
        warnings=warnings,
    )


def _row_score(row: list[Any], following_rows: list[list[Any]]) -> float:
    non_empty = [value for value in row if value not in (None, "")]
    if not non_empty:
        return -1.0
    text_ratio = sum(isinstance(value, str) for value in non_empty) / len(non_empty)
    unique_ratio = len({str(value) for value in non_empty}) / len(non_empty)
    following_width = max(
        (
            sum(value not in (None, "") for value in candidate)
            for candidate in following_rows
        ),
        default=0,
    )
    width_ratio = min(1.0, len(non_empty) / max(1, following_width))
    return 0.5 * text_ratio + 0.3 * unique_ratio + 0.2 * width_ratio


def _non_blank_row_runs(rows: list[list[Any]]) -> list[tuple[int, int]]:
    """Return populated regions separated by at least two consecutive blank rows."""
    runs: list[tuple[int, int]] = []
    start: int | None = None
    last_populated: int | None = None
    blank_count = 0
    for index, row in enumerate(rows):
        populated = any(value not in (None, "") for value in row)
        if populated:
            if start is None:
                start = index
            last_populated = index
            blank_count = 0
        elif start is not None:
            blank_count += 1
            if blank_count >= 2:
                if last_populated is not None:
                    runs.append((start, last_populated))
                start = None
                last_populated = None
                blank_count = 0
    if start is not None and last_populated is not None:
        runs.append((start, last_populated))
    return runs


def _discover_xlsx(
    file_bytes: bytes,
    *,
    file_sha256: str,
    raw_file_id: uuid.UUID | None,
    original_filename: str | None,
) -> SourceCatalog:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    tables: list[SourceTable] = []
    catalog_warnings: list[ParseWarning] = []
    try:
        for ws in wb.worksheets:
            matrix = [list(row) for row in ws.iter_rows(values_only=True)]
            runs = _non_blank_row_runs(matrix)
            if not runs:
                catalog_warnings.append(
                    ParseWarning(
                        code="empty_sheet",
                        message=f"Worksheet {ws.title!r} contains no table data.",
                        severity="info",
                        location=ws.title,
                    )
                )
                continue
            for region_number, (start, end) in enumerate(runs, start=1):
                region_rows = matrix[start : end + 1]
                populated_cols = [
                    index
                    for index in range(
                        max((len(row) for row in region_rows), default=0)
                    )
                    if any(
                        index < len(row) and row[index] not in (None, "")
                        for row in region_rows
                    )
                ]
                if not populated_cols:
                    continue
                if len(region_rows) == 1 and len(populated_cols) == 1:
                    catalog_warnings.append(
                        ParseWarning(
                            code="title_row_ignored",
                            message="Ignored an isolated worksheet title row.",
                            severity="info",
                            location=(
                                f"{ws.title}!"
                                f"{get_column_letter(populated_cols[0] + 1)}"
                                f"{start + 1}"
                            ),
                        )
                    )
                    continue
                min_col, max_col = min(populated_cols), max(populated_cols)
                region_rows = [
                    (row + [None] * (max_col + 1))[min_col : max_col + 1]
                    for row in region_rows
                ]
                candidates = min(10, len(region_rows))
                scores = [
                    _row_score(region_rows[index], region_rows[index + 1 : index + 6])
                    for index in range(candidates)
                ]
                header_offset = max(range(candidates), key=scores.__getitem__)
                confidence = max(0.4, min(0.99, scores[header_offset]))
                headers = region_rows[header_offset]
                repeated_headers = sum(
                    row == headers for row in region_rows[header_offset + 1 :]
                )
                data_rows = [
                    row
                    for row in region_rows[header_offset + 1 :]
                    if any(value not in (None, "") for value in row) and row != headers
                ]
                min_row = start + header_offset + 1
                max_row = end + 1
                cell_range = (
                    f"{get_column_letter(min_col + 1)}{min_row}:"
                    f"{get_column_letter(max_col + 1)}{max_row}"
                )
                table_warnings: list[ParseWarning] = []
                if header_offset:
                    table_warnings.append(
                        ParseWarning(
                            code="preamble_detected",
                            message=f"Detected {header_offset} preamble/title row(s).",
                            location=f"{ws.title}!{cell_range}",
                        )
                    )
                if repeated_headers:
                    table_warnings.append(
                        ParseWarning(
                            code="repeated_header",
                            message=(
                                f"Removed {repeated_headers} repeated header row(s)."
                            ),
                            location=f"{ws.title}!{cell_range}",
                        )
                    )
                merged = [
                    str(merged_range)
                    for merged_range in ws.merged_cells.ranges
                    if _ranges_overlap(cell_range, str(merged_range))
                ]
                if merged:
                    table_warnings.append(
                        ParseWarning(
                            code="merged_cells",
                            message=(
                                "Merged cells occur within the discovered table region."
                            ),
                            location=f"{ws.title}!{cell_range}",
                            details={"ranges": merged},
                        )
                    )
                display_name = ws.title
                if len(runs) > 1:
                    display_name = f"{ws.title} table {region_number}"
                tables.append(
                    _profile_table(
                        file_sha256=file_sha256,
                        raw_file_id=raw_file_id,
                        original_filename=original_filename,
                        display_name=display_name,
                        location=SourceLocation(
                            sheet_name=ws.title,
                            cell_range=cell_range,
                            header_row=min_row,
                            data_start_row=min_row + 1,
                        ),
                        headers=headers,
                        rows=data_rows,
                        confidence=confidence,
                        warnings=table_warnings,
                    )
                )
    finally:
        wb.close()
    return SourceCatalog(
        raw_file_id=raw_file_id,
        file_sha256=file_sha256,
        original_filename=original_filename,
        file_type="xlsx",
        tables=tables,
        warnings=catalog_warnings,
    )


def _ranges_overlap(left: str, right: str) -> bool:
    left_bounds = cast(tuple[int, int, int, int], range_boundaries(left))
    right_bounds = cast(tuple[int, int, int, int], range_boundaries(right))
    left_min_col, left_min_row, left_max_col, left_max_row = left_bounds
    right_min_col, right_min_row, right_max_col, right_max_row = right_bounds
    return not (
        left_max_col < right_min_col
        or right_max_col < left_min_col
        or left_max_row < right_min_row
        or right_max_row < left_min_row
    )


def discover_source_tables(
    file_bytes: bytes,
    file_type: str,
    *,
    raw_file_id: uuid.UUID | None = None,
    file_sha256: str | None = None,
    original_filename: str | None = None,
) -> SourceCatalog:
    """Discover and profile every tabular region in a CSV or XLSX source."""
    digest = file_sha256 or hashlib.sha256(file_bytes).hexdigest()
    if file_type == "csv":
        return _discover_csv(
            file_bytes,
            file_sha256=digest,
            raw_file_id=raw_file_id,
            original_filename=original_filename,
        )
    if file_type == "xlsx":
        return _discover_xlsx(
            file_bytes,
            file_sha256=digest,
            raw_file_id=raw_file_id,
            original_filename=original_filename,
        )
    raise ValueError(f"Source catalog discovery does not support {file_type!r}")


def load_workbook_from_bytes(file_bytes: bytes) -> Workbook:
    """Load an Excel workbook from bytes for read-only inspection."""
    return load_workbook(
        filename=io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )


def get_sheet_names(file_bytes: bytes) -> list[str]:
    """Return the list of sheet names in an Excel workbook."""
    wb = load_workbook_from_bytes(file_bytes)
    names = list(wb.sheetnames)
    wb.close()
    return names


def summarise_sheet(
    file_bytes: bytes,
    *,
    sheet_name: str | None = None,
    max_sample_values: int = 8,
    max_distinct_categorical: int = 15,
    max_cols: int = 52,
) -> dict[str, Any]:
    """Build a smart per-column summary of a spreadsheet sheet."""
    wb = load_workbook_from_bytes(file_bytes)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws = cast(Worksheet, wb[sheet_name])
    else:
        ws = cast(Worksheet, wb.active)
        sheet_name = ws.title

    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0
    col_count = min(max_cols, total_cols) if total_cols else max_cols

    all_rows: list[list[Any]] = []
    for worksheet_row in ws.iter_rows(max_col=col_count, values_only=True):
        all_rows.append(list(worksheet_row))

    wb.close()

    if not all_rows:
        return {
            "sheet_name": sheet_name,
            "total_rows": 0,
            "total_cols": total_cols,
            "header_row": 0,
            "columns": [],
        }

    header_row_idx = 0
    for idx, candidate_row in enumerate(all_rows):
        non_empty = [c for c in candidate_row if c is not None]
        if non_empty and all(isinstance(c, str) for c in non_empty):
            header_row_idx = idx
            break
        if non_empty:
            header_row_idx = idx
            break

    headers = (
        all_rows[header_row_idx]
        if header_row_idx < len(all_rows)
        else [None] * col_count
    )
    data_rows = all_rows[header_row_idx + 1 :]

    column_summaries: list[dict[str, Any]] = []
    for col_idx in range(col_count):
        col_letter = get_column_letter(col_idx + 1)
        header_val = headers[col_idx] if col_idx < len(headers) else None

        values = []
        for data_row in data_rows:
            if col_idx < len(data_row) and data_row[col_idx] is not None:
                values.append(data_row[col_idx])

        if not values:
            column_summaries.append(
                {
                    "column_letter": col_letter,
                    "header": str(header_val) if header_val is not None else None,
                    "non_empty_count": 0,
                    "first_values": [],
                    "last_values": [],
                    "dominant_type": None,
                    "type_inconsistencies": None,
                    "distinct_values": None,
                }
            )
            continue

        first_vals = [str(v) for v in values[:max_sample_values]]
        last_vals = (
            [str(v) for v in values[-max_sample_values:]]
            if len(values) > max_sample_values
            else []
        )

        type_counts = Counter(type(v).__name__ for v in values)
        dominant_type = type_counts.most_common(1)[0][0]
        type_inconsistencies = None
        if len(type_counts) > 1:
            minority_types = {
                t: c for t, c in type_counts.items() if t != dominant_type
            }
            type_inconsistencies = (
                f"Mostly {dominant_type} ({type_counts[dominant_type]}/"
                f"{len(values)}) but also: "
                f"{', '.join(f'{t} ({c})' for t, c in minority_types.items())}"
            )

        distinct_raw = {str(v) for v in values}
        distinct_values = None
        if len(distinct_raw) <= max_distinct_categorical:
            distinct_values = sorted(distinct_raw)

        column_summaries.append(
            {
                "column_letter": col_letter,
                "header": str(header_val) if header_val is not None else None,
                "non_empty_count": len(values),
                "first_values": first_vals,
                "last_values": last_vals,
                "dominant_type": dominant_type,
                "type_inconsistencies": type_inconsistencies,
                "distinct_values": distinct_values,
            }
        )

    return {
        "sheet_name": sheet_name,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "header_row": header_row_idx + 1,
        "data_start_row": header_row_idx + 2,
        "columns": column_summaries,
    }


def read_csv_to_polars(file_bytes: bytes, **options: Any) -> pl.DataFrame:
    """Parse CSV bytes into a Polars DataFrame."""
    return pl.read_csv(io.BytesIO(file_bytes), **options)


def read_excel_to_polars(
    file_bytes: bytes,
    sheet_name: str | None = None,
    **options: Any,
) -> pl.DataFrame:
    """Parse Excel bytes into a Polars DataFrame."""
    return cast(
        pl.DataFrame,
        pl.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=sheet_name,
            engine="openpyxl",
            **options,
        ),
    )


def parse_pdf_to_text(file_bytes: bytes) -> dict[str, Any]:
    """Extract text and page references from a PDF."""
    try:
        from pypdf import PdfReader
    except ImportError:  # pragma: no cover
        return {"pages": [], "full_text": "", "error": "pypdf not installed"}

    reader = PdfReader(io.BytesIO(file_bytes))
    pages = []
    full_text_parts = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        pages.append({"page": i, "text": text})
        full_text_parts.append(text)
    return {
        "pages": pages,
        "full_text": "\n".join(full_text_parts),
    }


def parse_email_to_dict(file_bytes: bytes) -> dict[str, Any]:
    """Parse an email message into structured headers and body."""
    msg = email.message_from_bytes(file_bytes)
    headers = dict(msg.items())
    body_text_parts = []
    body_html_parts = []
    attachments = []

    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            filename = part.get_filename()
            if filename:
                attachments.append(filename)
                continue
            payload = part.get_payload(decode=True)
            if payload is None:
                continue
            if not isinstance(payload, bytes):
                continue
            text = payload.decode("utf-8", errors="ignore")
            if content_type == "text/plain":
                body_text_parts.append(text)
            elif content_type == "text/html":
                body_html_parts.append(text)
    else:
        payload = msg.get_payload(decode=True)
        if isinstance(payload, bytes):
            body_text_parts.append(payload.decode("utf-8", errors="ignore"))

    return {
        "headers": headers,
        "body_text": "\n".join(body_text_parts),
        "body_html": "\n".join(body_html_parts),
        "attachments": attachments,
    }


def parse_text_document(file_bytes: bytes, mime_type: str) -> dict[str, Any]:
    """Parse plain-text, markdown, or Word documents into text."""
    text = file_bytes.decode("utf-8", errors="ignore")
    if "wordprocessingml" in mime_type or file_bytes.startswith(b"PK"):
        try:
            from docx import Document
        except ImportError:  # pragma: no cover
            return {"text": "", "error": "python-docx not installed"}
        doc = Document(io.BytesIO(file_bytes))
        text = "\n".join(p.text for p in doc.paragraphs)
    return {"text": text, "sections": []}


def extract_evidence_chunks(
    parsed_document: dict[str, Any],
    raw_file_id: str,
    chunk_size: int = 1000,
    chunk_overlap: int = 200,
) -> list[dict[str, Any]]:
    """Split a parsed document into searchable evidence chunks."""
    text = ""
    if "full_text" in parsed_document:
        text = parsed_document["full_text"]
    elif "text" in parsed_document:
        text = parsed_document["text"]
    elif "body_text" in parsed_document:
        text = parsed_document["body_text"]

    chunks: list[dict[str, Any]] = []
    start = 0
    idx = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if not chunk.strip():
            break
        chunks.append(
            {
                "raw_file_id": raw_file_id,
                "evidence_type": "text_chunk",
                "content": chunk,
                "page_ref": None,
                "chunk_index": idx,
                "metadata": {},
            }
        )
        start = end - chunk_overlap
        idx += 1
    return chunks


def build_polars_from_mapping_source(
    file_bytes: bytes,
    file_type: str,
    source_table: str,
) -> pl.DataFrame:
    """Load any supported raw file into a Polars DataFrame for a mapping source."""
    if file_type == "csv":
        return read_csv_to_polars(file_bytes)
    if file_type == "xlsx":
        return read_excel_to_polars(file_bytes, sheet_name=source_table or None)
    raise ValueError(f"Cannot build Polars DataFrame from file type '{file_type}'")


def profile_polars_dataframe(
    df: pl.DataFrame,
    source_table: str,
) -> dict[str, Any]:
    """Profile a Polars DataFrame and emit column-level statistics."""
    columns = []
    for name in df.columns:
        series = df[name]
        non_null = series.drop_nulls()
        sample_values = [str(v) for v in non_null.head(5).to_list()]
        distinct = non_null.unique().to_list()
        columns.append(
            {
                "column": name,
                "dtype": str(series.dtype),
                "null_count": int(series.null_count()),
                "non_null_count": len(non_null),
                "unique_count": len(distinct),
                "sample_values": sample_values,
            }
        )
    return {
        "source_table": source_table,
        "row_count": len(df),
        "columns": columns,
    }
