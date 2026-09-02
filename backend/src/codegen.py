"""Deterministic artifact generation and Polars execution pipeline.

This module is the single-file Polars pipeline for the target transformation
layer. From an approved mapping specification it generates a standalone Polars
transformation script and a machine-readable mapping file. It also executes the
Polars pipeline, runs validation checks, captures results, and records lineage
back to source files and mapping columns.

All outputs are shaped by the supplied TargetSchema.
"""

from __future__ import annotations

import ast
import json
import re
import subprocess
import sys
import tempfile
import textwrap
import uuid
from datetime import UTC
from pathlib import Path
from typing import Any

from config import get_settings
from db_ops import create_generated_artifact, get_raw_file_by_id
from file_ops import ObjectStore
from mapping_specs import load_mapping_spec, load_target_schema_from_spec
from models import (
    GeneratedArtifact,
    GeneratedPipelineScript,
    MappingFile,
    PipelineOutputFolder,
    TargetSchema,
)

# ---------------------------------------------------------------------------
# Source-loading helpers (inlined into every generated pipeline.py)
# ---------------------------------------------------------------------------

_SOURCE_LOADING_HELPERS = """\
def load_mapping_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _catalog_tables(mapping_spec: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        table
        for catalog in mapping_spec.get("source_catalogs", [])
        for table in catalog.get("tables", [])
    ]


def _resolve_source_path(source_folder: Path, table: dict[str, Any]) -> Path:
    original_filename = table.get("original_filename")
    raw_file_id = table.get("raw_file_id")
    candidates = []
    if original_filename:
        exact = source_folder / original_filename
        if exact.is_file():
            candidates.append(exact)
    if raw_file_id:
        candidates.extend(
            p
            for p in source_folder.glob(f"{raw_file_id}_*")
            if p.is_file() and p not in candidates
        )
    if len(candidates) != 1:
        raise ValueError(
            f"Expected exactly one local file for source table "
            f"{table['source_table_id']}, found {len(candidates)}"
        )
    path = candidates[0]
    expected_hash = table.get("file_sha256")
    if expected_hash:
        actual_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        if actual_hash != expected_hash:
            raise ValueError(
                f"Content hash mismatch for source table {table['source_table_id']}"
            )
    return path


def _catalog_column_names(table: dict[str, Any]) -> list[str]:
    return [column["normalized_name"] for column in table.get("columns", [])]


def _rows_to_dataframe(
    rows: list[list[Any]],
    table: dict[str, Any],
) -> pl.DataFrame:
    names = _catalog_column_names(table)
    width = len(names)
    header = [(str(value).strip() if value is not None else "") for value in rows[0]]
    data_rows = []
    for row in rows[1:]:
        padded = (list(row) + [None] * width)[:width]
        comparable = [
            str(value).strip() if value is not None else "" for value in padded
        ]
        if not any(value not in (None, "") for value in padded):
            continue
        if comparable == header[:width]:
            continue
        data_rows.append(padded)
    return pl.DataFrame(data_rows, schema=names, orient="row", infer_schema_length=None)


def _load_csv_table(path: Path, table: dict[str, Any]) -> pl.DataFrame:
    location = table["location"]
    encoding = location.get("encoding") or "utf-8"
    skip_rows = int(location["header_row"]) - 1
    kwargs: dict[str, Any] = {
        "skip_rows": skip_rows,
        "encoding": encoding,
        "truncate_ragged_lines": True,
        "infer_schema_length": 1000,
    }
    delimiter = location.get("delimiter")
    if delimiter:
        kwargs["separator"] = delimiter
    quote_char = location.get("quote_char")
    if quote_char:
        kwargs["quote_char"] = quote_char
    df = pl.read_csv(path, **kwargs)
    names = _catalog_column_names(table)
    if len(df.columns) == len(names):
        df = df.rename({old: new for old, new in zip(df.columns, names)})
    return df


def _load_xlsx_table(path: Path, table: dict[str, Any]) -> pl.DataFrame:
    location = table["location"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[location["sheet_name"]]
        min_col, min_row, max_col, max_row = range_boundaries(
            location["cell_range"]
        )
        rows = [
            list(row)
            for row in worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        ]
    finally:
        workbook.close()
    return _rows_to_dataframe(rows, table)


def _load_legacy_source_dataframes(
    source_folder: Path,
) -> dict[str, pl.DataFrame]:
    result = {}
    for path in sorted(source_folder.iterdir()):
        if path.suffix.lower() == ".csv":
            result[path.stem] = pl.read_csv(path)
        elif path.suffix.lower() in {".xlsx", ".xls"}:
            sheets = pl.read_excel(path, sheet_id=None, engine="openpyxl")
            if isinstance(sheets, dict):
                result.update({str(name): frame for name, frame in sheets.items()})
            else:
                result[path.stem] = sheets
    return result


def load_source_dataframes(
    mapping_spec: dict[str, Any],
    source_folder: Path,
) -> dict[str, pl.DataFrame]:
    tables = _catalog_tables(mapping_spec)
    if not tables:
        return _load_legacy_source_dataframes(source_folder)
    result = {}
    for table in tables:
        path = _resolve_source_path(source_folder, table)
        file_type = path.suffix.lower()
        if file_type == ".csv":
            frame = _load_csv_table(path, table)
        elif file_type == ".xlsx":
            frame = _load_xlsx_table(path, table)
        else:
            raise ValueError(
                f"Unsupported tabular source type for {table['source_table_id']}: "
                f"{file_type}"
            )
        result[table["source_table_id"]] = frame
    return result
"""

# ---------------------------------------------------------------------------
# Expression helpers (inlined into every generated pipeline.py)
# ---------------------------------------------------------------------------

_EXPRESSION_HELPERS = """\
from polars import col, when, coalesce, lit, concat_str

null = None
Int64 = pl.Int64
Float64 = pl.Float64
String = pl.String
Date = pl.Date
Datetime = pl.Datetime
Boolean = pl.Boolean


def concat(*args: Any) -> pl.Expr:
    return pl.concat_str([
        pl.lit(x) if not isinstance(x, pl.Expr) else x for x in args
    ])
"""

# ---------------------------------------------------------------------------
# LLM codegen
# ---------------------------------------------------------------------------


def _gather_codegen_context(spec_id: uuid.UUID) -> dict[str, Any]:
    """Gather catalogs, evidence, rules, and file summaries for codegen."""
    from db_ops import (
        get_mapping_spec,
        get_session,
        get_spreadsheet_profile,
        search_evidence_by_text,
    )
    from mapping import _gather_targeted_evidence
    from models import BusinessRule, RawFile
    from sqlmodel import select

    with get_session() as session:
        spec = get_mapping_spec(session, spec_id)
        if spec is None:
            raise ValueError(f"Mapping spec not found: {spec_id}")

        mapping_spec = load_mapping_spec(spec_id)
        target_schema = load_target_schema_from_spec(mapping_spec)

        raw_file_records = [
            session.get(RawFile, rid) for rid in spec.source_raw_file_ids
        ]
        raw_files: list[RawFile] = [rf for rf in raw_file_records if rf is not None]

        source_catalogs: list[dict[str, Any]] = []
        for raw_file in raw_files:
            profile = get_spreadsheet_profile(session, raw_file.id)
            if profile:
                source_catalogs.append(profile.profile_json)

        raw_file_summary = [
            {
                "filename": rf.original_filename,
                "mime_type": rf.mime_type,
                "raw_file_id": str(rf.id),
            }
            for rf in raw_files
        ]

        evidence_items = _gather_targeted_evidence(
            session,
            spec.client_id,
            target_schema,
            source_catalogs,
            search_evidence_by_text=search_evidence_by_text,
            top_k_per_query=5,
            max_total=40,
        )
        business_rules = list(
            session.exec(
                select(BusinessRule).where(
                    BusinessRule.client_id == spec.client_id,
                    BusinessRule.status == "approved",
                )
            ).all()
        )

    return {
        "source_catalogs": source_catalogs,
        "evidence_items": evidence_items,
        "business_rules": business_rules,
        "raw_file_summary": raw_file_summary,
        "mapping_spec": mapping_spec,
        "target_schema": target_schema,
    }


def _extract_python_code(text: str) -> str:
    """Strip markdown fences and leading/trailing whitespace from LLM output."""
    code = text.strip()
    if code.startswith("```python"):
        code = code[len("```python") :]
    elif code.startswith("```"):
        code = code[3:]
    if code.endswith("```"):
        code = code[:-3]
    return code.strip()


def _validate_generated_code(code: str) -> None:
    """Validate generated pipeline code structure.

    Raises ``RuntimeError`` if the code is invalid.
    """
    try:
        ast.parse(code)
    except SyntaxError as exc:
        raise RuntimeError(f"LLM returned invalid Python: {exc}") from exc
    if "--source-folder" not in code:
        raise RuntimeError("Generated pipeline does not handle --source-folder")
    if "--output-folder" not in code:
        raise RuntimeError("Generated pipeline does not handle --output-folder")


_EXPRESSION_FIELDS = (  # noqa: E501
    "polars_expression",
    "filter_expression",
    "aggregation_expression",
)


def _projection_for_codegen(col: dict[str, Any]) -> dict[str, Any]:
    """Project a mapping column for the codegen prompt.

    Mappings are prose-first: Polars expression fields are stripped so the
    model works from plain-English ``transformation_logic`` plus structured
    parameters. For legacy specs where the prose is empty but an expression
    was stored, the expression is inlined as read-only reference text.
    """
    projected = {
        key: value for key, value in col.items() if key not in _EXPRESSION_FIELDS
    }
    if not (col.get("transformation_logic") or "").strip():
        legacy = next(
            (col.get(field) for field in _EXPRESSION_FIELDS if col.get(field)),
            None,
        )
        if legacy:
            projected["legacy_polars_reference"] = (
                f"Legacy implementation for reference only — reimplement from "
                f"this intent: {legacy}"
            )
    return projected


def _normalize_generated_pipeline(code: str) -> str:
    """Apply deterministic fixes to LLM-generated pipeline code.

    Mirrors the safe subset of ``_normalize_polars_expression`` at whole-file
    level. Most importantly, bare string literals passed to ``.then()`` /
    ``.otherwise()`` are wrapped in ``pl.lit()`` — Polars would otherwise
    interpret them as column names.
    """
    # LLMs often use pandas/string-style title casing or strip names.
    code = code.replace(".str.title()", ".str.to_titlecase()")
    code = code.replace(".str.strip()", ".str.strip_chars()")
    # Bare string literals in then()/otherwise() are column references in
    # Polars; wrap them as literals. Already-wrapped calls don't match.
    code = re.sub(
        r"\.then\((['\"])([^'\"]*)\1\)",
        r".then(pl.lit(\1\2\1))",
        code,
    )
    code = re.sub(
        r"\.otherwise\((['\"])([^'\"]*)\1\)",
        r".otherwise(pl.lit(\1\2\1))",
        code,
    )
    return code


def _codegen_with_context(
    spec_id: uuid.UUID,
    base_code: str,
    error_message: str | None = None,
    focus_column: str | None = None,
) -> str:
    """Generate pipeline.py via LLM using mapping context + a base code template."""
    from mapping import build_codegen_prompt, call_codegen_llm

    context = _gather_codegen_context(spec_id)
    columns = context["mapping_spec"].get("columns", [])
    if focus_column is not None:
        columns = [
            col for col in columns if col.get("target_column") == focus_column
        ] or context["mapping_spec"].get("columns", [])
    mapping_json = json.dumps(
        [_projection_for_codegen(col) for col in columns], indent=2, default=str
    )
    messages = build_codegen_prompt(
        context["target_schema"],
        context["source_catalogs"],
        context["evidence_items"],
        context["business_rules"],
        context["raw_file_summary"],
        mapping_json,
        base_code,
        error_message=error_message,
        focus_column=focus_column,
    )
    settings = get_settings()
    code = call_codegen_llm(
        messages,
        model=settings.codegen_model,
        api_key=settings.openai_api_key,
        base_url=settings.openai_base_url,
    )
    code = _extract_python_code(code)
    code = _normalize_generated_pipeline(code)
    _validate_generated_code(code)
    return code


def generate_pipeline_code(
    mapping_spec: dict[str, Any],
    target_schema: TargetSchema,
    spec_id: uuid.UUID | None = None,
) -> str:
    """Generate pipeline code via LLM, using the deterministic draft as template.

    When *spec_id* is ``None`` the deterministic draft is returned directly.
    """
    draft = generate_polars_script(mapping_spec, target_schema)
    if spec_id is not None:
        return _codegen_with_context(spec_id, draft)
    return draft


# ---------------------------------------------------------------------------
# Artifact generation
# ---------------------------------------------------------------------------


def _transformation_comment(col: dict[str, Any]) -> str:
    """Build the English intent comment embedded above a placeholder block."""
    target_column = col["target_column"]
    logic = (col.get("transformation_logic") or "").strip()
    lines = [f"# TARGET COLUMN: {target_column}"]
    if logic:
        lines.append(f"# Transformation: {logic}")
    return "\n".join(lines) + "\n"


def _generate_transform_code(col: dict[str, Any], source_key: str) -> str:
    """Generate Python code for a single mapping column's transformation.

    Mappings are prose-first: expression/filter/aggregation logic is plain
    English in ``transformation_logic``, so the deterministic draft emits a
    syntactically valid placeholder plus the English intent comment. The LLM
    codegen pass replaces placeholders with real implementations. Lookups are
    fully structured and generated deterministically.
    """
    target_column = col["target_column"]
    ttype = col.get("transformation_type", "expression")

    if ttype == "filter":
        return (
            _transformation_comment(col)
            + "# Transformation type: filter\n"
            + "# PLACEHOLDER: implement this filter\n"
        )

    if ttype == "lookup":
        lookup_table = col.get("lookup_source_table", "")
        lookup_key = col.get("lookup_key", "")
        lookup_value = col.get("lookup_value", "")
        left_key = (
            col["source_columns"][0]["source_column"]
            if col.get("source_columns")
            else lookup_key
        )
        return textwrap.dedent(f"""\
            _lut = (
                source_dfs[{lookup_table!r}]
                .select([{lookup_key!r}, {lookup_value!r}])
                .unique()
            )
            df = df.join(
                _lut,
                left_on={left_key!r},
                right_on={lookup_key!r},
                how="left",
            ).rename({{{lookup_value!r}: {target_column!r}}})
        """)

    if ttype == "aggregation":
        agg_table = col.get("aggregation_source_table", "")
        agg_key = col.get("aggregation_group_key", "")
        base_key = (
            col["source_columns"][0]["source_column"]
            if col.get("source_columns")
            else agg_key
        )
        lines = [
            _transformation_comment(col).rstrip("\n"),
            f"# Transformation type: aggregation from {agg_table!r} "
            f"grouped by {agg_key!r}",
            "# PLACEHOLDER: implement this aggregation",
            f"_agg_df = source_dfs[{agg_table!r}]",
            f"_grouped = _agg_df.group_by({agg_key!r}).agg(",
            f"    pl.col({base_key!r}).alias({base_key!r}),",
            ")",
            "df = df.join(",
            "    _grouped,",
            f"    left_on={base_key!r},",
            f"    right_on={agg_key!r},",
            '    how="left",',
            f").rename({{{base_key!r}: {target_column!r}}})",
        ]
        return "\n".join(lines) + "\n"

    # Default: expression
    source_columns = col.get("source_columns", [])

    if not source_columns:
        return f"df = df.with_columns(pl.lit(None).alias({target_column!r}))\n"

    first_source = source_columns[0]["source_column"]

    return (
        _transformation_comment(col) + f"df = df.with_columns("
        f"pl.col({first_source!r}).alias({target_column!r}))"
        f"  # PLACEHOLDER: implement the transformation above\n"
    )


def generate_polars_script(
    mapping_spec: dict[str, Any],
    target_schema: TargetSchema,
) -> str:
    """Generate a unique standalone Polars pipeline script from a mapping spec.

    Each generated script embeds the transformation logic for every column,
    eliminating the need for eval() on LLM-generated expressions at runtime.
    The source-loading helpers are inlined so the script is fully self-contained.
    """
    imports = """\
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
"""

    # Group columns by target table
    columns_by_table: dict[str, list[dict[str, Any]]] = {}
    for c in mapping_spec["columns"]:
        columns_by_table.setdefault(c["target_table"], []).append(c)

    # Determine base source key per target table
    base_keys: dict[str, str] = {}
    for target_table, columns in columns_by_table.items():
        base_source = target_table
        for c in columns:
            if c.get("source_columns"):
                source_ref = c["source_columns"][0]
                base_source = source_ref.get(
                    "source_table_id",
                    source_ref.get("source_table", target_table),
                )
                break
        base_keys[target_table] = base_source

    # Generate transformation code per target table
    table_blocks: list[str] = []
    for target_table in sorted(columns_by_table):
        columns = columns_by_table[target_table]
        source_key = base_keys[target_table]
        lines = [
            f"# --- target table: {target_table} ---",
            f"_sk = {source_key!r}",
            "if _sk not in source_dfs:",
            "    _stem = Path(_sk).stem",
            "    _sk = _stem if _stem in source_dfs else _sk",
            "df = source_dfs[_sk].clone()",
            "",
        ]

        # Filters first
        for c in columns:
            if c.get("transformation_type") == "filter":
                lines.append(_generate_transform_code(c, source_key))

        # Remaining transformations
        for c in columns:
            if c.get("transformation_type") != "filter":
                lines.append(_generate_transform_code(c, source_key))

        target_cols = [c["target_column"] for c in columns]
        col_set = {cc["target_column"] for cc in columns}
        selected = [repr(c) for c in target_cols if c in col_set]
        lines.append(f"df = df.select([{', '.join(selected)}])")
        lines.append(f"_write_table({target_table!r}, df)")
        lines.append("")

        table_blocks.append("\n".join(lines))

    # Enforce dtypes
    dtype_blocks: list[str] = []
    for table in target_schema.tables:
        casts = []
        for col in table.columns:
            if col.dtype:
                casts.append(f'("{{name}}", "{col.dtype}")')
        if casts:
            dtype_blocks.append(
                f'    "tables": {{"{table.name}": [{", ".join(casts)}]}},'
            )

    # Assemble full script
    main_body = textwrap.dedent("""\
        def _write_table(name: str, df: pl.DataFrame) -> None:
            _dtype = _DTYPES.get(name, {})
            for col_name, dtype_str in _dtype.items():
                if col_name not in df.columns:
                    continue
                current = str(df[col_name].dtype).lower()
                if "null" in current:
                    continue
                if dtype_str == "Date":
                    if "date" in current and "datetime" not in current:
                        continue
                    if "datetime" in current:
                        df = df.with_columns(
                            pl.col(col_name).dt.date().alias(col_name)
                        )
                    else:
                        df = df.with_columns(
                            pl.col(col_name)
                            .str.to_date(strict=False)
                            .alias(col_name)
                        )
                elif dtype_str == "Datetime":
                    if "datetime" in current:
                        continue
                    df = df.with_columns(
                        pl.col(col_name)
                        .str.to_datetime(strict=False)
                        .alias(col_name)
                    )
                else:
                    polars_dtype = getattr(pl, dtype_str, None)
                    if polars_dtype is not None:
                        df = df.with_columns(
                            pl.col(col_name)
                            .cast(polars_dtype, strict=False)
                        )
            df.write_csv(_OUTPUT_DIR / f"{name}.csv")


        def build_target_tables() -> None:
    """)
    for block in table_blocks:
        main_body += textwrap.indent(block, "    ") + "\n"

    main_body += textwrap.dedent("""\

        if __name__ == "__main__":
            parser = argparse.ArgumentParser()
            parser.add_argument("--source-folder", required=True, type=Path)
            parser.add_argument("--output-folder", default=".", type=Path)
            args = parser.parse_args()

            pipeline_dir = Path(__file__).parent
            mapping_spec = load_mapping_json(pipeline_dir / "mapping.json")
            source_dfs = load_source_dataframes(mapping_spec, args.source_folder)
            _OUTPUT_DIR = args.output_folder
            _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            build_target_tables()
    """)

    dtype_dict: dict[str, dict[str, str]] = {}
    for table in target_schema.tables:
        for col in table.columns:
            if col.dtype:
                dtype_dict.setdefault(table.name, {})[col.name] = col.dtype

    script = "\n".join(
        [
            '"""Auto-generated single-file Polars transformation pipeline."""',
            imports,
            _SOURCE_LOADING_HELPERS,
            _EXPRESSION_HELPERS,
            f"_DTYPES: dict[str, dict[str, str]] = {json.dumps(dtype_dict, indent=4)}",
            "",
            "_OUTPUT_DIR: Path = Path('.')",
            "",
            main_body,
        ]
    )

    return script


def generate_polars_pipeline_script(
    mapping_spec: dict[str, Any],
    target_schema: TargetSchema,
    *,
    spec_id: uuid.UUID | None = None,
) -> GeneratedPipelineScript:
    """Generate a standalone single-file Polars pipeline script."""
    target_tables = sorted({c["target_table"] for c in mapping_spec["columns"]})
    content = generate_pipeline_code(mapping_spec, target_schema, spec_id)
    return GeneratedPipelineScript(
        file_path=Path("pipeline.py"),
        content=content,
        target_tables=target_tables,
    )


def generate_mapping_json(
    mapping_spec: dict[str, Any],
) -> MappingFile:
    """Generate a human- and machine-readable mapping.json file."""
    return MappingFile(
        file_path=Path("mapping.json"),
        content=mapping_spec,
    )


def generate_artifact_set(
    spec_id: uuid.UUID,
    output_folder: str | Path,
) -> list[GeneratedArtifact]:
    """Generate the full set of artifacts for an approved mapping spec."""
    from db_ops import get_session

    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    mapping_spec = load_mapping_spec(spec_id)
    target_schema = load_target_schema_from_spec(mapping_spec)

    artifacts: list[GeneratedArtifact] = []
    with get_session() as session:
        # single-file Polars pipeline
        pipeline = generate_polars_pipeline_script(
            mapping_spec, target_schema, spec_id=spec_id
        )
        pipeline_path = output_folder / "pipeline.py"
        pipeline_path.write_text(pipeline.content)
        artifacts.append(
            create_generated_artifact(
                session,
                spec_id,
                "pipeline",
                str(pipeline_path.relative_to(output_folder)),
                pipeline.content,
                [],
            )
        )

        # mapping.json
        mapping_file = generate_mapping_json(mapping_spec)
        mapping_path = output_folder / "mapping.json"
        mapping_path.write_text(json.dumps(mapping_file.content, indent=2, default=str))
        artifacts.append(
            create_generated_artifact(
                session,
                spec_id,
                "mapping",
                str(mapping_path.relative_to(output_folder)),
                mapping_path.read_text(),
                [],
            )
        )

    return artifacts


def generate_output_folder(
    spec_id: uuid.UUID,
    output_folder: str | Path,
    object_store: ObjectStore,
) -> PipelineOutputFolder:
    """Generate the complete client deliverable folder.

    Writes pipeline.py, mapping.json, and one CSV per target table into
    output_folder.
    """
    from datetime import datetime

    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    mapping_spec = load_mapping_spec(spec_id)
    target_schema = load_target_schema_from_spec(mapping_spec)

    # Write mapping.json
    mapping_file = generate_mapping_json(mapping_spec)
    mapping_path = output_folder / "mapping.json"
    mapping_path.write_text(json.dumps(mapping_file.content, indent=2, default=str))

    # Write pipeline.py
    pipeline = generate_polars_pipeline_script(
        mapping_spec, target_schema, spec_id=spec_id
    )
    pipeline_path = output_folder / "pipeline.py"
    pipeline_path.write_text(pipeline.content)

    # Execute pipeline to produce results.csv, repairing automatically on failure.
    _execute_with_repair(pipeline_path, output_folder, object_store, spec_id)

    # Keep the generated artifacts plus one CSV per target table.
    allowed = {"pipeline.py", "mapping.json"}
    target_tables = {f"{c['target_table']}.csv" for c in mapping_spec["columns"]}
    for path in output_folder.iterdir():
        if (
            path.is_file()
            and path.name not in allowed
            and path.name not in target_tables
        ):
            path.unlink()

    first_table = next(iter({c["target_table"] for c in mapping_spec["columns"]}), None)
    if first_table:
        results_csv_path = output_folder / f"{first_table}.csv"
    else:
        results_csv_path = output_folder / "results.csv"
    return PipelineOutputFolder(
        folder_path=output_folder,
        pipeline_py_path=output_folder / "pipeline.py",
        mapping_json_path=output_folder / "mapping.json",
        results_csv_path=results_csv_path,
        generated_at=datetime.now(UTC),
    )


_MAX_PIPELINE_REPAIRS = 2


def _extract_focus_column(code: str, error_message: str) -> str | None:
    """Best-effort extraction of the target column behind a pipeline failure.

    Walks up from the failing traceback line to the nearest ``# TARGET
    COLUMN`` comment emitted by the draft; falls back to an ``alias(...)``
    call near the failure.
    """
    matches = re.findall(r'File "[^"]*pipeline\.py", line (\d+)', error_message)
    if not matches:
        return None
    lines = code.splitlines()
    if not lines:
        return None
    line_no = min(int(matches[-1]), len(lines))
    for idx in range(line_no - 1, max(-1, line_no - 41), -1):
        match = re.search(r"#\s*TARGET COLUMN:\s*(.+)", lines[idx])
        if match:
            return match.group(1).strip()
    window = "\n".join(lines[max(0, line_no - 6) : min(len(lines), line_no + 5)])
    match = re.search(r"\.alias\(\s*['\"]([^'\"]+)['\"]", window)
    if match:
        return match.group(1)
    return None


def _execute_with_repair(
    pipeline_py_path: Path,
    output_folder: Path,
    object_store: ObjectStore,
    spec_id: uuid.UUID,
) -> dict[str, Path]:
    """Execute the generated pipeline, regenerating it automatically on failure.

    Up to ``_MAX_PIPELINE_REPAIRS`` LLM repairs are attempted with the runtime
    error fed back, focused on the failing column when it can be identified.
    The last execution error is re-raised once repairs are exhausted.
    """
    last_error: RuntimeError | None = None
    for attempt in range(_MAX_PIPELINE_REPAIRS + 1):
        try:
            return execute_generated_pipeline(
                pipeline_py_path, output_folder, object_store, spec_id
            )
        except RuntimeError as exc:
            last_error = exc
        if attempt >= _MAX_PIPELINE_REPAIRS:
            break
        failed_code = pipeline_py_path.read_text()
        focus_column = _extract_focus_column(failed_code, str(last_error))
        corrected = _codegen_with_context(
            spec_id,
            failed_code,
            error_message=str(last_error),
            focus_column=focus_column,
        )
        pipeline_py_path.write_text(corrected)
    assert last_error is not None
    raise last_error


def execute_generated_pipeline(
    pipeline_py_path: str | Path,
    output_folder: str | Path,
    object_store: ObjectStore,
    spec_id: uuid.UUID,
) -> dict[str, Path]:
    """Run a generated pipeline.py script and capture its CSV outputs."""
    from db_ops import get_mapping_spec, get_session

    pipeline_py_path = Path(pipeline_py_path)
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmpdir:
        source_folder = Path(tmpdir) / "sources"
        source_folder.mkdir(parents=True, exist_ok=True)

        with get_session() as session:
            spec = get_mapping_spec(session, spec_id)
            if spec is None:
                raise ValueError(f"Mapping spec not found: {spec_id}")
            for raw_file_id in spec.source_raw_file_ids:
                raw_file = get_raw_file_by_id(session, raw_file_id)
                if raw_file is None:
                    continue
                data = object_store.get(raw_file.storage_key)
                runtime_name = f"{raw_file.id}_{raw_file.original_filename}"
                (source_folder / runtime_name).write_bytes(data)

        result = subprocess.run(
            [
                sys.executable,
                str(pipeline_py_path),
                "--source-folder",
                str(source_folder),
                "--output-folder",
                str(output_folder),
            ],
            capture_output=True,
            timeout=300,
        )
        if result.returncode != 0:
            stderr = result.stderr.decode(errors="replace")
            raise RuntimeError(
                f"Pipeline execution failed (exit {result.returncode}):\n{stderr}"
            )

    csv_files = sorted(output_folder.glob("*.csv"))
    return {p.stem: p for p in csv_files}
